# -*- coding: utf-8 -*-
"""
Created on Tue Nov 15 15:30:21 2022

@author: Kaitlin
Objective: Using the functions in the Storygraph Extraction file, update an excel spreadsheet of book titles with
the desired information about the book including the storygraph link
Note: This does not account for different editions or formats.
"""

import openpyxl
import Storygraph_Extraction_v3 as SE
# The below packages are imported through the SE module, so they don't need to be imported again
# import requests
# from bs4 import BeautifulSoup

def apply_details(book, col_ref): #pulls the main details and puts them into the excel spreadsheet
    link = SE.get_SGLink(book) # getting the storygraph link
    row[col_ref].value = link 
    page = requests.get(link) 
    soup = BeautifulSoup(page.content, 'html.parser') 
    # Grabs book genre, mood, pacing, page count, publication date, and series name if available and saves to appropriate columns
    # If you change any columns after the link, you will need to change the column values here
    # Could use the prompt list to find the value, but this way saves test time and space if you're not changing this order
    book_type, genre = SE.get_genre(soup)
    row[col_ref+1].value = book_type
    row[col_ref+2].value = genre
    row[col_ref+3].value = SE.get_mood(soup)
    row[col_ref+4].value = SE.get_pacing(soup)
    row[col_ref+5].value = SE.get_book_len(soup)
    row[col_ref+6].value = SE.get_book_date(soup)
    row[2].value = SE.get_series(soup) 
    return soup # Saving to use for author name 



dataframe = openpyxl.load_workbook('Book Backlog.xlsx') # Opening our excel spreadsheet to read and write to
dataframe1 = dataframe.active

i=0
prompt = []
# Saving a list of the column values and finding the column for storygraph links
for col in dataframe1.iter_cols(1, 20):
    prompt.append(col[0].value)
    if col[0].value == "Storygraph Link":
        col_ref = i
    i += 1

for row in dataframe1.iter_rows(10,1000): # Determines how far down the spreadsheet to check
    
    if row[col_ref].value == None: # Checks if there's no link
        if row[0].value != None: # Checks if there's a title
            if row[1].value == None: # Checks for author name - will grab it if it's not available
                book = row[0].value
                soup = apply_details(book, col_ref) 
                row[1].value = SE.get_author(soup)
            else: # Adds author name to search - helps if it's a book with a common title
                book = row[0].value
                book += " (" + row[1].value + ")"
                print(book)
                soup = apply_details(book, col_ref) 

dataframe.save('Book Backlog.xlsx') # Saves the spreadsheet